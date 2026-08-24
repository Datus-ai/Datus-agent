# Copyright 2025-present DatusAI, Inc.
# Licensed under the Apache License, Version 2.0.
# See http://www.apache.org/licenses/LICENSE-2.0 for details.

"""Spreadsheet loading through :mod:`datus.tools.db_tools.data_file_loader`.

Integration rather than unit tier because ``read_xlsx`` lives in DuckDB's
autoloadable ``excel`` extension: a real external artifact, fetched from
extensions.duckdb.org unless the deployment baked it into the image. The
``excel_extension`` fixture installs it and lets any failure surface — a skip
here would hide the exact breakage this feature is most likely to hit in
production (an egress-less pod with no baked extension).
"""

from __future__ import annotations

import duckdb
import openpyxl
import pytest
import xlwt

from datus.tools.db_tools.data_file_loader import (
    DataFileError,
    convert_legacy_xls,
    inspect_file,
    load_file,
    ownership_tag,
    registered_objects,
)

pytestmark = pytest.mark.acceptance


@pytest.fixture
def connection(tmp_path):
    con = duckdb.connect(str(tmp_path / "catalog.duckdb"))
    try:
        yield con
    finally:
        con.close()


@pytest.fixture
def excel_connection(connection):
    """A connection with the ``excel`` extension loaded.

    ``LOAD`` first, and in CI that is the only step: the workflow installs the
    extension once up front and points
    ``DATUS_DUCKDB_EXTENSION_DIRECTORY`` at it, so these tests reach no network
    — which is the same posture as the runtime image, where the extension is
    baked for pods with no egress.

    ``INSTALL`` remains as the fallback for a developer machine that has never
    fetched it. It is not wrapped in a skip: an environment where the extension
    can be neither loaded nor fetched should fail loudly, because that is
    precisely the breakage this feature is most likely to hit in production.
    """
    try:
        connection.execute("LOAD excel")
    except duckdb.IOException:
        connection.execute("INSTALL excel")
        connection.execute("LOAD excel")
    return connection


def _write_workbook(path, *, with_title_rows=True):
    """A workbook shaped like real business exports: a title, a spacer, then the
    header — plus a second data sheet and an empty one."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Detail"
    if with_title_rows:
        sheet.append(["2024 Sales Detail"])
        sheet.append([])
    sheet.append(["region", "qty", "amount"])
    for row in [("east", 3, 10.5), ("west", 5, 22.0), ("east", 2, 7.25)]:
        sheet.append(list(row))
    summary = workbook.create_sheet("Summary")
    summary.append(["region", "total"])
    summary.append(["east", 17.75])
    workbook.create_sheet("Notes")
    workbook.save(path)
    return path


class TestSheetDiscovery:
    def test_one_table_per_readable_sheet(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "jeffshop.xlsx")

        loaded, skipped = load_file(source, "jeffshop.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)

        assert {item.sheet for item in loaded} == {"Detail", "Summary"}
        assert [item.sheet for item in skipped] == ["Notes"]

    def test_table_names_derive_from_file_and_sheet(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "jeffshop.xlsx")
        loaded, _ = load_file(source, "jeffshop.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)
        assert sorted(item.table for item in loaded) == ["jeffshop_detail", "jeffshop_summary"]

    def test_empty_sheet_is_skipped_not_fatal(self, tmp_path, excel_connection):
        """A blank tab is normal in real workbooks and must not cost the others."""
        source = _write_workbook(tmp_path / "s.xlsx")
        loaded, skipped = load_file(source, "s.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)
        assert len(loaded) == 2
        assert [item.sheet for item in skipped] == ["Notes"]

    def test_workbook_with_only_empty_sheets_raises(self, tmp_path, excel_connection):
        workbook = openpyxl.Workbook()
        workbook.active.title = "Blank"
        source = tmp_path / "empty.xlsx"
        workbook.save(source)
        with pytest.raises(DataFileError):
            load_file(source, "empty.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)

    def test_single_sheet_selection(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")
        loaded, _ = load_file(
            source, "s.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path, sheet="Summary"
        )
        assert [item.sheet for item in loaded] == ["Summary"]

    def test_unknown_sheet_names_the_available_ones(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")
        with pytest.raises(DataFileError) as excinfo:
            load_file(source, "s.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path, sheet="Nope")
        assert "Detail" in str(excinfo.value)


class TestHeaderAndRange:
    def test_header_is_found_below_title_rows(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")
        loaded, _ = load_file(source, "s.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)

        detail = next(item for item in loaded if item.sheet == "Detail")
        assert detail.header_row == 3
        assert detail.preview_columns == ["region", "qty", "amount"]

    def test_row_count_excludes_title_rows(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")
        loaded, _ = load_file(source, "s.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)
        detail = next(item for item in loaded if item.sheet == "Detail")
        assert detail.row_count == 3

    def test_range_is_bounded_to_the_used_extent(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")
        loaded, _ = load_file(source, "s.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)
        detail = next(item for item in loaded if item.sheet == "Detail")
        assert detail.used_range == "A3:C6"

    def test_no_all_null_padding_rows(self, tmp_path, excel_connection):
        """Regression: ``read_xlsx`` needs a complete range, and padding its end
        reads every blank cell as a NULL *row* — inflating counts, driving the
        null rate to ~98%, and adding a phantom NULL group to every aggregate.
        The model cannot tell that apart from genuinely sparse data."""
        source = _write_workbook(tmp_path / "s.xlsx")
        load_file(source, "s.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)

        groups = excel_connection.execute('SELECT region, count(*) FROM "s_detail" GROUP BY 1 ORDER BY 1').fetchall()
        assert groups == [("east", 2), ("west", 1)]

    def test_null_percentage_is_zero_for_dense_data(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")
        loaded, _ = load_file(source, "s.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)
        detail = next(item for item in loaded if item.sheet == "Detail")
        assert [float(column["null_percentage"]) for column in detail.columns] == [0.0, 0.0, 0.0]

    def test_explicit_header_row_overrides_detection(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")
        loaded, _ = load_file(
            source,
            "s.xlsx",
            connection=excel_connection,
            conversion_cache_dir=tmp_path,
            sheet="Detail",
            header_row=1,
        )
        assert loaded[0].header_row == 1
        assert loaded[0].preview_columns[0] == "2024 Sales Detail"

    def test_header_row_past_the_end_fails_with_the_reason(self, tmp_path, excel_connection):
        """Nothing loadable is an error, not a silent empty result — the caller
        passed a header_row and needs to know it was wrong."""
        source = _write_workbook(tmp_path / "s.xlsx")
        with pytest.raises(DataFileError) as excinfo:
            load_file(
                source,
                "s.xlsx",
                connection=excel_connection,
                conversion_cache_dir=tmp_path,
                sheet="Detail",
                header_row=99,
            )
        assert "header row is past the end" in str(excinfo.value)

    def test_a_bad_header_row_does_not_leave_a_stale_table(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")
        with pytest.raises(DataFileError):
            load_file(
                source,
                "s.xlsx",
                connection=excel_connection,
                conversion_cache_dir=tmp_path,
                sheet="Detail",
                header_row=99,
            )
        assert registered_objects(excel_connection) == {}

    def test_header_in_the_first_row_needs_no_offset(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx", with_title_rows=False)
        loaded, _ = load_file(
            source, "s.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path, sheet="Detail"
        )
        assert loaded[0].header_row == 1
        assert loaded[0].row_count == 3


class TestFreshness:
    def test_edits_within_the_recorded_range_are_seen_without_a_reload(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")
        load_file(source, "s.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path, sheet="Detail")
        assert excel_connection.execute('SELECT sum(amount) FROM "s_detail"').fetchone()[0] == pytest.approx(39.75)

        workbook = openpyxl.load_workbook(source)
        workbook["Detail"]["C4"] = 1000.0
        workbook.save(source)

        assert excel_connection.execute('SELECT sum(amount) FROM "s_detail"').fetchone()[0] == pytest.approx(1029.25)

    def test_rows_appended_past_the_range_need_a_reload(self, tmp_path, excel_connection):
        """The pinned range is what keeps padding rows out; the cost is that a
        grown sheet needs re-registering. This is why the tool tells the model to
        load before analysing."""
        source = _write_workbook(tmp_path / "s.xlsx")
        kwargs = dict(connection=excel_connection, conversion_cache_dir=tmp_path, sheet="Detail")
        load_file(source, "s.xlsx", **kwargs)

        workbook = openpyxl.load_workbook(source)
        workbook["Detail"].append(["north", 9, 100.0])
        workbook.save(source)

        assert excel_connection.execute('SELECT count(*) FROM "s_detail"').fetchone()[0] == 3

        load_file(source, "s.xlsx", existing_objects=registered_objects(excel_connection), **kwargs)
        assert excel_connection.execute('SELECT count(*) FROM "s_detail"').fetchone()[0] == 4


class TestIdempotency:
    def test_repeated_loads_do_not_grow_the_catalog(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "jeffshop.xlsx")
        kwargs = dict(connection=excel_connection, conversion_cache_dir=tmp_path)

        first, _ = load_file(source, "jeffshop.xlsx", existing_objects=registered_objects(excel_connection), **kwargs)
        second, _ = load_file(source, "jeffshop.xlsx", existing_objects=registered_objects(excel_connection), **kwargs)
        third, _ = load_file(source, "jeffshop.xlsx", existing_objects=registered_objects(excel_connection), **kwargs)

        names = sorted(item.table for item in first)
        assert sorted(item.table for item in second) == names
        assert sorted(item.table for item in third) == names
        assert sorted(registered_objects(excel_connection)) == names

    def test_ownership_comment_identifies_the_sheet(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")
        load_file(source, "s.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)

        catalog = registered_objects(excel_connection)
        assert catalog["s_detail"] == ownership_tag("s.xlsx", "Detail")
        assert catalog["s_summary"] == ownership_tag("s.xlsx", "Summary")


class TestNonAsciiNames:
    def test_cjk_sheet_names_yield_ascii_table_names(self, tmp_path, excel_connection):
        """Instruction from the product side: table names stay ``[a-z0-9_]`` so
        the model never has to quote them."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Q3明细"
        sheet.append(["region", "amount"])
        sheet.append(["east", 1.5])
        second = workbook.create_sheet("汇总")
        second.append(["region", "total"])
        second.append(["east", 1.5])
        source = tmp_path / "jeffshop.xlsx"
        workbook.save(source)

        loaded, _ = load_file(source, "jeffshop.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)

        names = sorted(item.table for item in loaded)
        assert names == ["jeffshop_q3", "jeffshop_s2"]
        for name in names:
            assert name.replace("_", "").isalnum()

    def test_original_sheet_name_is_reported_back(self, tmp_path, excel_connection):
        """The sanitised name loses meaning, so the mapping has to travel in the
        result for the model to read."""
        workbook = openpyxl.Workbook()
        workbook.active.title = "汇总"
        workbook.active.append(["a"])
        workbook.active.append([1])
        source = tmp_path / "book.xlsx"
        workbook.save(source)

        loaded, _ = load_file(source, "book.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)

        assert loaded[0].sheet == "汇总"
        assert loaded[0].table == "book_s1"

    def test_cjk_column_names_are_preserved_and_queryable(self, tmp_path, excel_connection):
        """Columns keep their original names: unlike a table name, a column's
        meaning cannot survive transliteration, and the model needs it."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "S"
        sheet.append(["品类", "金额"])
        sheet.append(["鞋类", 299.0])
        source = tmp_path / "book.xlsx"
        workbook.save(source)

        loaded, _ = load_file(source, "book.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)

        assert loaded[0].preview_columns == ["品类", "金额"]
        total = excel_connection.execute('SELECT sum("金额") FROM "book_s"').fetchone()[0]
        assert total == pytest.approx(299.0)


class TestLegacyXls:
    def _write_xls(self, path):
        book = xlwt.Workbook()
        sheet = book.add_sheet("S")
        for column, name in enumerate(["a", "b"]):
            sheet.write(0, column, name)
        sheet.write(1, 0, 1)
        sheet.write(1, 1, 2)
        book.save(str(path))
        return path

    def test_loads_through_a_parquet_conversion(self, tmp_path, connection):
        """No excel extension involved: DuckDB cannot read BIFF at all."""
        source = self._write_xls(tmp_path / "old.xls")

        loaded, _ = load_file(source, "old.xls", connection=connection, conversion_cache_dir=tmp_path / "conv")

        assert loaded[0].row_count == 1
        assert loaded[0].preview_columns == ["a", "b"]

    def test_conversion_is_cached(self, tmp_path):
        source = self._write_xls(tmp_path / "old.xls")
        cache = tmp_path / "conv"

        first = convert_legacy_xls(source, cache)
        second = convert_legacy_xls(source, cache)

        assert first == second
        assert first.suffix == ".parquet"

    def test_conversion_is_redone_when_the_source_changes(self, tmp_path):
        """The one format that needs an explicit staleness check, because it is
        the only one not read lazily."""
        source = self._write_xls(tmp_path / "old.xls")
        cache = tmp_path / "conv"
        first = convert_legacy_xls(source, cache)

        book = xlwt.Workbook()
        sheet = book.add_sheet("S")
        sheet.write(0, 0, "a")
        sheet.write(1, 0, 1)
        sheet.write(2, 0, 2)
        book.save(str(source))

        second = convert_legacy_xls(source, cache)
        assert second != first
        assert not first.exists(), "the superseded conversion should be cleaned up"


class TestInspectSpreadsheet:
    def test_returns_the_unparsed_grid(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")

        details = inspect_file(source, connection=excel_connection, sheet="Detail")

        assert details["sheets"] == ["Detail", "Summary", "Notes"]
        detail = details["sheet_details"][0]
        assert detail["detected_header_row"] == 3
        # The point of inspect: the title row is still visible, so a wrong header
        # guess can be diagnosed and corrected via header_row.
        assert detail["raw_rows"][0][0] == "2024 Sales Detail"
        assert detail["used_rows"] == 6
        assert detail["used_columns"] == 3

    def test_creates_nothing(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")
        inspect_file(source, connection=excel_connection)
        assert registered_objects(excel_connection) == {}

    def test_lists_every_sheet_when_none_is_named(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")
        details = inspect_file(source, connection=excel_connection)
        assert [entry["sheet"] for entry in details["sheet_details"]] == ["Detail", "Summary", "Notes"]

    def test_unknown_sheet_raises(self, tmp_path, excel_connection):
        source = _write_workbook(tmp_path / "s.xlsx")
        with pytest.raises(DataFileError):
            inspect_file(source, connection=excel_connection, sheet="Nope")


class TestExtentComesFromValuesNotFormatting:
    """``worksheet.max_row`` reports the sheet *dimension*, which counts any cell
    carrying only formatting. Trusting it reintroduces exactly the padding bug the
    exact range exists to prevent, from a cause no author would think to look for:
    someone bolded a blank cell.
    """

    def _sheet_with_styled_blank(self, path, blank_cell="B200"):
        from openpyxl.styles import Font

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "S"
        sheet.append(["region", "qty"])
        sheet.append(["east", 3])
        sheet.append(["west", 5])
        sheet[blank_cell].font = Font(bold=True)
        workbook.save(path)
        return path

    def test_styled_blank_cell_does_not_inflate_the_range(self, tmp_path, excel_connection):
        source = self._sheet_with_styled_blank(tmp_path / "styled.xlsx")
        loaded, _ = load_file(source, "styled.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)
        assert loaded[0].used_range == "A1:B3"

    def test_row_count_is_the_real_row_count(self, tmp_path, excel_connection):
        source = self._sheet_with_styled_blank(tmp_path / "styled.xlsx")
        loaded, _ = load_file(source, "styled.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)
        assert loaded[0].row_count == 2

    def test_no_phantom_null_group(self, tmp_path, excel_connection):
        source = self._sheet_with_styled_blank(tmp_path / "styled.xlsx")
        load_file(source, "styled.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)
        groups = excel_connection.execute('SELECT region, count(*) FROM "styled_s" GROUP BY 1 ORDER BY 1').fetchall()
        assert groups == [("east", 1), ("west", 1)]

    def test_styled_blank_column_does_not_add_phantom_columns(self, tmp_path, excel_connection):
        source = self._sheet_with_styled_blank(tmp_path / "styled.xlsx", blank_cell="E5")
        loaded, _ = load_file(source, "styled.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)
        assert loaded[0].preview_columns == ["region", "qty"]

    def test_interior_blank_row_is_not_truncated(self, tmp_path, excel_connection):
        """The other half of the trade-off: an exact range must not be paired with
        DuckDB's inferred stop_at_empty, or a separator row silently drops every
        group below it."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "S"
        sheet.append(["region", "qty"])
        sheet.append(["east", 1])
        sheet.append([])
        sheet.append(["west", 2])
        sheet.append(["north", 3])
        source = tmp_path / "gap.xlsx"
        workbook.save(source)

        loaded, _ = load_file(source, "gap.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)

        regions = excel_connection.execute('SELECT region FROM "gap_s" WHERE region IS NOT NULL ORDER BY 1').fetchall()
        assert regions == [("east",), ("north",), ("west",)]


class TestFormulaCellsCountTowardsExtent:
    """The mirror of the padding bug, and just as silent.

    ``data_only=True`` yields the *cached* value of a formula cell, and a
    workbook written by a library rather than by Excel has no cached values at
    all. Reading bounds that way makes a trailing formula row look empty, so the
    range stops short and the rows are dropped with nothing to indicate it.
    """

    def _sheet_with_uncached_formula(self, path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "S"
        sheet.append(["region", "qty"])
        sheet.append(["east", 3])
        sheet.append(["west", 5])
        # openpyxl writes the formula but no cached result, exactly like any
        # spreadsheet produced without Excel ever opening it.
        sheet["A4"] = "=A2"
        sheet["B4"] = "=SUM(B2:B3)"
        workbook.save(path)
        return path

    def test_trailing_formula_row_is_inside_the_range(self, tmp_path, excel_connection):
        source = self._sheet_with_uncached_formula(tmp_path / "f.xlsx")
        loaded, _ = load_file(source, "f.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)
        assert loaded[0].used_range == "A1:B4"

    def test_the_formula_row_is_not_dropped(self, tmp_path, excel_connection):
        source = self._sheet_with_uncached_formula(tmp_path / "f.xlsx")
        loaded, _ = load_file(source, "f.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)
        assert loaded[0].row_count == 3

    def test_a_formula_only_column_still_counts(self, tmp_path, excel_connection):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "S"
        sheet.append(["a", "b", "total"])
        sheet.append([1, 2, None])
        sheet["C2"] = "=A2+B2"
        source = tmp_path / "g.xlsx"
        workbook.save(source)

        loaded, _ = load_file(source, "g.xlsx", connection=excel_connection, conversion_cache_dir=tmp_path)

        assert loaded[0].preview_columns == ["a", "b", "total"]


class TestLegacyXlsSheets:
    """A ``.xls`` has sheets like any other workbook.

    Treating the file as one unit made ``sheet`` a silent no-op, because
    ``pandas.read_excel`` defaults to the first worksheet — a caller asking for
    the second one received the first one's data, which is worse than an error.
    """

    def _multi_sheet_xls(self, path):
        book = xlwt.Workbook()
        first = book.add_sheet("First")
        first.write(0, 0, "a")
        first.write(1, 0, 1)
        second = book.add_sheet("Second")
        second.write(0, 0, "b")
        second.write(1, 0, 99)
        book.save(str(path))
        return path

    def test_every_sheet_becomes_its_own_table(self, tmp_path, connection):
        source = self._multi_sheet_xls(tmp_path / "multi.xls")

        loaded, _ = load_file(source, "multi.xls", connection=connection, conversion_cache_dir=tmp_path)

        assert sorted(item.sheet for item in loaded) == ["First", "Second"]

    def test_requesting_a_sheet_returns_that_sheet(self, tmp_path, connection):
        source = self._multi_sheet_xls(tmp_path / "multi.xls")

        loaded, _ = load_file(source, "multi.xls", connection=connection, conversion_cache_dir=tmp_path, sheet="Second")

        assert [item.sheet for item in loaded] == ["Second"]
        assert loaded[0].preview_columns == ["b"]
        assert connection.execute(f'SELECT * FROM "{loaded[0].table}"').fetchall() == [(99,)]

    def test_unknown_sheet_names_the_available_ones(self, tmp_path, connection):
        source = self._multi_sheet_xls(tmp_path / "multi.xls")
        with pytest.raises(DataFileError) as excinfo:
            load_file(source, "multi.xls", connection=connection, conversion_cache_dir=tmp_path, sheet="Nope")
        assert "First" in str(excinfo.value)

    def test_each_sheet_caches_its_own_conversion(self, tmp_path, connection):
        """One cache entry per sheet — keying on the file alone would have the
        second sheet overwrite the first."""
        source = self._multi_sheet_xls(tmp_path / "multi.xls")
        cache = tmp_path / "conv"

        load_file(source, "multi.xls", connection=connection, conversion_cache_dir=cache)

        assert len(list(cache.glob("*.parquet"))) == 2
