# Copyright 2025-present DatusAI, Inc.
# Licensed under the Apache License, Version 2.0.
# See http://www.apache.org/licenses/LICENSE-2.0 for details.

"""Expose uploaded data files (spreadsheets, CSV, Parquet, JSON) as SQL tables.

Every supported format is surfaced the same way: a **lazy DuckDB VIEW** over the
file on disk. Nothing is copied, so there is no second source of truth to keep
in sync, and re-running a load is idempotent. That is the whole reason this
module exists instead of a materialise-and-track-staleness design: no content
hashes, no registry, no refresh command.

How current a view is, precisely, differs by format:

* CSV / Parquet / JSON — fully lazy. The view names only the path, so every
  query re-reads the file and always sees the latest bytes.
* Spreadsheets — lazy *within a pinned range*. ``read_xlsx`` demands a complete
  ``range`` and the end of it cannot be padded (see :func:`sheet_bounds`), so
  the range is fixed at load time. Edits to existing cells show up immediately;
  rows appended past the recorded end do not, until the file is loaded again.

Hence the standing advice to the model: call the load before analysing. It is
idempotent and near-free when nothing changed, which makes "just re-load" a
cheaper rule than any staleness check.

Format support, and why it differs:

* ``.csv`` / ``.tsv``  — ``read_csv_auto``, DuckDB core.
* ``.parquet``         — ``read_parquet``, statically linked into the wheel.
* ``.json`` / ``.jsonl`` / ``.ndjson`` — ``read_json_auto``, statically linked.
* ``.xlsx`` / ``.xlsm`` — ``read_xlsx`` from the autoloadable ``excel``
  extension. Reading sheet *contents* is all the extension does: it cannot list
  sheet names and does not report the used range, so openpyxl supplies both
  (see :func:`enumerate_sheets` / :func:`sheet_bounds` for why each matters).
* ``.xls``             — DuckDB has no support at all (``read_xlsx`` is strictly
  OOXML/zip). Converted once to Parquet via pandas+xlrd, then treated as
  Parquet, so everything downstream of "make it queryable" stays uniform.

Deliberately unsupported: ``.xlsb``, ``.ods``. Both would need another optional
pandas engine, and a clear "convert to .xlsx" error is more useful than a
half-working path.
"""

from __future__ import annotations

import hashlib
import os
import re
import tempfile
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from datus.utils.loggings import get_logger

logger = get_logger(__name__)

#: Datasource key the SaaS control plane registers this catalog under. Must stay
#: in sync with ``datus_backend.utils.fs_isolation.LOCAL_FILES_DATASOURCE`` —
#: ``tests/unit/test_local_files_datasource.py`` asserts the two agree.
LOCAL_FILES_DATASOURCE = "local_files"

#: Rows of the preview handed back to the model. Enough to see the shape of the
#: data and spot a misread header; small enough not to dominate the turn.
PREVIEW_ROW_LIMIT = 20

#: How far down a sheet to look for the header row. Real-world spreadsheets put
#: a title, a blank line and maybe a note above it; beyond this it is not a
#: header, it is a layout, and the caller should pass ``header_row`` explicitly.
HEADER_PROBE_ROW_LIMIT = 30

#: Columns to sample when probing for the header row.
HEADER_PROBE_COLUMN_LIMIT = 64

_CSV_SUFFIXES = frozenset({".csv", ".tsv"})

#: Text encodings DuckDB's CSV reader accepts (verified against 1.5.2 — it
#: rejects anything else outright, ``gbk`` included, so a detected label has to
#: be mapped onto this set rather than passed through).
_DUCKDB_CSV_ENCODINGS = frozenset({"utf-8", "utf-16", "latin-1", "gb18030", "big5", "shift_jis", "cp1252"})

#: What a detected label maps to. ``gb18030`` is a superset of GBK and GB2312, so
#: the whole family lands there; the rest collapse onto DuckDB's spelling.
#: Order to try when the detected label is unusable, restricted to encodings
#: whose DuckDB reader was verified to round-trip a CJK/accented sample intact.
#: Frequency-ordered for the files this tool actually sees: Excel on a Chinese
#: Windows first. ``latin-1`` is deliberately absent — it decodes any byte
#: sequence, so offering it would mask every real answer behind mojibake.
_ENCODING_CANDIDATES = ("gb18030", "shift_jis", "cp1252")

#: Below this many bytes a detected encoding is not solid enough to refuse a file
#: over — the detector will label a handful of bytes as almost anything. Set just
#: past "too short to hold a CSV row at all"; a real header plus one row clears
#: it comfortably and is still refused when it genuinely is Big5.
_MIN_SAMPLE_FOR_REFUSAL = 16

#: Accepted by DuckDB's reader but corrupted by it: a Big5 trailing byte may fall
#: in the ASCII range (``0x42`` here), and the reader loses the row boundary
#: after it — ``金額\n台北`` comes back as one field. Verified on 1.5.2 against a
#: file Python decodes correctly, so this is the reader, not the detection.
#: Detected rather than silently attempted, because mojibake with no error is the
#: one outcome worse than a refusal.
_DUCKDB_BROKEN_CSV_ENCODINGS = frozenset({"big5"})

_ENCODING_ALIASES = {
    "utf_8": "utf-8",
    "utf8": "utf-8",
    "ascii": "utf-8",
    "utf_16": "utf-16",
    "utf_16_le": "utf-16",
    "utf_16_be": "utf-16",
    "gbk": "gb18030",
    "gb2312": "gb18030",
    "gb18030": "gb18030",
    "cp936": "gb18030",
    "cp949": "gb18030",
    "euc_kr": "gb18030",
    "big5": "big5",
    "big5hkscs": "big5",
    "shift_jis": "shift_jis",
    "sjis": "shift_jis",
    "cp932": "shift_jis",
    "cp1252": "cp1252",
    "windows_1252": "cp1252",
    "latin_1": "latin-1",
    "iso8859_1": "latin-1",
}
_PARQUET_SUFFIXES = frozenset({".parquet", ".pq"})
_JSON_SUFFIXES = frozenset({".json", ".jsonl", ".ndjson"})
_EXCEL_SUFFIXES = frozenset({".xlsx", ".xlsm"})
_LEGACY_EXCEL_SUFFIXES = frozenset({".xls"})
_UNSUPPORTED_SUFFIXES = frozenset({".xlsb", ".ods"})

SUPPORTED_SUFFIXES = frozenset(
    _CSV_SUFFIXES | _PARQUET_SUFFIXES | _JSON_SUFFIXES | _EXCEL_SUFFIXES | _LEGACY_EXCEL_SUFFIXES
)

#: Table functions that read arbitrary paths. Harmless when *we* emit them
#: inside a VIEW definition (the path was already validated against the
#: filesystem policy); a hole when they appear in model-authored SQL, because
#: they would read any file the process can see and bypass the path policy
#: entirely. :func:`find_file_reading_functions` is the gate.
FILE_READING_FUNCTIONS = frozenset(
    {
        "read_csv",
        "read_csv_auto",
        "read_parquet",
        "read_json",
        "read_json_auto",
        "read_ndjson",
        "read_ndjson_auto",
        "read_xlsx",
        "read_text",
        "read_blob",
        "parquet_scan",
        "csv_scan",
        "glob",
    }
)

_FUNCTION_CALL_RE = re.compile(r"\b([a-z_][a-z0-9_]*)\s*\(", re.IGNORECASE)

_IDENT_SAFE_RE = re.compile(r"[^A-Za-z0-9]+")


class DataFileError(Exception):
    """A load failed for a reason the model can act on (bad format, empty sheet)."""


@dataclass
class LoadedTable:
    """One VIEW created by a load, plus what the model needs to query it."""

    table: str
    source_file: str
    sheet: Optional[str] = None
    header_row: Optional[int] = None
    used_range: Optional[str] = None
    encoding: Optional[str] = None
    row_count: int = 0
    columns: List[Dict[str, Any]] = field(default_factory=list)
    preview_columns: List[str] = field(default_factory=list)
    preview_rows: List[List[Any]] = field(default_factory=list)
    materialized: bool = False

    def to_dict(self) -> Dict[str, Any]:
        payload: Dict[str, Any] = {
            "table": self.table,
            "source_file": self.source_file,
            "row_count": self.row_count,
            "columns": self.columns,
            "preview_columns": self.preview_columns,
            "preview_rows": self.preview_rows,
        }
        if self.sheet is not None:
            payload["sheet"] = self.sheet
        if self.header_row is not None:
            payload["header_row"] = self.header_row
        if self.used_range is not None:
            payload["used_range"] = self.used_range
        if self.encoding is not None:
            payload["encoding"] = self.encoding
        if self.materialized:
            payload["materialized"] = True
        return payload


@dataclass
class SkippedSheet:
    sheet: str
    reason: str

    def to_dict(self) -> Dict[str, str]:
        return {"sheet": self.sheet, "reason": self.reason}


# ---------------------------------------------------------------- identifiers


def sanitize_identifier(text: str) -> str:
    """Reduce arbitrary text to ``[a-z0-9_]``, or ``""`` when nothing survives.

    NFKD first so that accented latin (``é`` → ``e``) and full-width digits
    survive as their ASCII equivalents instead of being stripped. CJK has no
    such decomposition and legitimately reduces to empty — callers fall back to
    a positional or hashed name rather than transliterating, because
    romanisation is ambiguous (multi-reading characters) and a plausible-looking
    but wrong name is worse than an honest opaque one. The human-readable
    mapping travels in the tool result instead.
    """
    normalized = unicodedata.normalize("NFKD", text)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    cleaned = _IDENT_SAFE_RE.sub("_", ascii_only).strip("_").lower()
    return cleaned


def _short_hash(*parts: str) -> str:
    digest = hashlib.sha1("\x1f".join(parts).encode("utf-8")).hexdigest()
    return digest[:8]


def build_table_name(
    *,
    rel_path: str,
    sheet: Optional[str],
    sheet_index: int,
    taken: Sequence[str],
) -> str:
    """Deterministic, SQL-safe view name for one (file, sheet) pair.

    Deterministic matters: the same upload re-loaded in a later session must
    resolve to the same name, otherwise ``CREATE OR REPLACE`` stops being
    idempotent and the catalog accumulates duplicates.
    """
    stem = sanitize_identifier(Path(rel_path).stem) or f"t_{_short_hash(rel_path)}"
    if stem[0].isdigit():
        stem = f"t_{stem}"

    if sheet is None:
        candidate = stem
    else:
        sheet_part = sanitize_identifier(sheet) or f"s{sheet_index + 1}"
        if sheet_part[0].isdigit():
            sheet_part = f"s{sheet_part}"
        candidate = f"{stem}_{sheet_part}"

    if candidate not in taken:
        return candidate
    # Two distinct sources sanitizing to the same name. Suffix with a hash of
    # the real identity so the collision resolves the same way every time.
    return f"{candidate}_{_short_hash(rel_path, sheet or '')[:6]}"


def quote_identifier(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def quote_literal(value: Any) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def column_letter(index: int) -> str:
    """1-based column index → spreadsheet letters (1 → A, 27 → AA)."""
    if index < 1:
        raise ValueError(f"column index must be >= 1, got {index}")
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


# ------------------------------------------------------------------ SQL guard


def find_file_reading_functions(sql: str) -> List[str]:
    """Names from :data:`FILE_READING_FUNCTIONS` that appear called in ``sql``.

    A regex rather than a parse on purpose: this runs on every statement, and
    the question is only "does a file-reading table function appear here". Over-
    matching (a column literally named ``glob(``) fails closed with a clear
    message; under-matching would be a policy bypass, so the bias is deliberate.
    """
    found: List[str] = []
    for match in _FUNCTION_CALL_RE.finditer(sql):
        name = match.group(1).lower()
        if name in FILE_READING_FUNCTIONS and name not in found:
            found.append(name)
    return found


def unresolved_table_references(sql: str, known: Sequence[str]) -> List[str]:
    """Table references in ``sql`` that do not name a registered catalog object.

    A whitelist, deliberately, because a blacklist of file-reading syntax cannot
    be completed. DuckDB's *replacement scan* reads a path written directly as a
    table — ``SELECT * FROM '/data/tenants/other/x.parquet'``, globs included —
    with no function call anywhere for a name check to catch (verified against
    1.5.2: reads the file, parses as a plain SELECT). Requiring every reference
    to resolve to something the catalog already holds closes that, closes every
    file-reading function at once, and closes whatever DuckDB adds next.

    sqlglot lowers a quoted path into an ``Identifier``, so paths and real table
    names are the same node type — the difference is only that a path is not in
    ``known``. CTE names defined in the same statement resolve too.

    Returns the offending reference names; empty means every reference resolved.
    An unparseable statement returns a sentinel so the caller can fail closed.
    """
    import sqlglot
    from sqlglot import exp

    try:
        tree = sqlglot.parse_one(sql, dialect="duckdb")
    except Exception:
        return ["<unparseable>"]
    if tree is None:
        return ["<unparseable>"]

    allowed = {name.lower() for name in known}
    allowed.update(cte.alias_or_name.lower() for cte in tree.find_all(exp.CTE))

    offenders: List[str] = []
    for table in tree.find_all(exp.Table):
        name = table.name
        if not name:
            # A table *function* (``read_csv_auto(...)``, ``duckdb_views()``)
            # carries its name on the inner node, not on the Table. Treating it
            # as unnamed would exempt exactly the constructs this is here to
            # catch, including any reader DuckDB adds later.
            inner = table.this
            name = getattr(inner, "name", "") or ""
            if not name:
                if "<unnamed table expression>" not in offenders:
                    offenders.append("<unnamed table expression>")
                continue
            name = f"{name}()"
        if name.lower().removesuffix("()") not in allowed and name not in offenders:
            offenders.append(name)
    return offenders


# ------------------------------------------------------------------- openpyxl


def enumerate_sheets(path: Path) -> List[str]:
    """Sheet names of an xlsx/xlsm workbook, in workbook order.

    DuckDB cannot do this: there is no ``xlsx_sheets()`` table function, and
    ``read_xlsx(..., sheet='nope')`` reports "Sheet not found" *without* listing
    what does exist. Multi-sheet workbooks are the common case, so enumeration
    is a hard requirement, not a nicety.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise DataFileError(
            "Reading .xlsx requires the 'openpyxl' package, which is not installed in this environment."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def sheet_bounds(path: Path, sheet: str) -> Tuple[int, int]:
    """``(max_row, max_column)`` of the last cell that actually holds a value.

    This is load-bearing for correctness, not an optimisation. ``read_xlsx``
    needs a complete ``range`` (``'A3'`` alone is rejected), and overshooting the
    end is silently wrong: every padding cell reads back as a NULL *row*, so a
    2-row sheet scanned as ``A1:E200`` reports 199 rows, a 99% null rate, two
    phantom columns and an extra all-NULL group in every ``GROUP BY``. The model
    cannot tell that apart from genuinely sparse data.

    Which is exactly why ``worksheet.max_row`` / ``max_column`` are not used:
    they report the sheet's *dimension*, which counts any cell carrying only
    formatting. One bolded blank cell far below the data — routine in a
    hand-maintained export — inflates the extent to that row (verified: a
    styled ``B200`` on a 2-row sheet yields exactly the 199-row reading above).
    Scanning values costs one streaming pass, against a file every query
    re-reads anyway.
    """
    from openpyxl import load_workbook

    # ``data_only=False`` so a formula cell counts even when the workbook carries
    # no cached value for it — a file written by a library rather than by Excel
    # has none. With ``data_only=True`` such a cell reads as ``None``, the row
    # falls outside the bounds, and the data is silently dropped: the mirror
    # image of the padding bug this function exists to prevent.
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if sheet not in workbook.sheetnames:
            raise DataFileError(f"Sheet {sheet!r} not found in {path.name}. Available: {workbook.sheetnames}")
        worksheet = workbook[sheet]
        last_row = 0
        last_column = 0
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            widest = 0
            for column_index, value in enumerate(row, start=1):
                if value is not None and str(value).strip() != "":
                    widest = column_index
            if widest:
                last_row = row_index
                last_column = max(last_column, widest)
        return last_row, last_column
    finally:
        workbook.close()


# --------------------------------------------------------------- header probe


def _probe_grid(connection: Any, path: Path, sheet: str, max_row: int, max_column: int) -> List[Tuple[Any, ...]]:
    """Raw upper-left cells of a sheet, with no header or type inference.

    ``all_varchar`` avoids type-inference failures while probing, and
    ``stop_at_empty=false`` is mandatory: it defaults to on, so a blank spacer
    row under the title truncates the probe to a single row and the real header
    is never seen.
    """
    rows = min(max_row, HEADER_PROBE_ROW_LIMIT)
    cols = min(max_column, HEADER_PROBE_COLUMN_LIMIT)
    if rows < 1 or cols < 1:
        return []
    probe_range = f"A1:{column_letter(cols)}{rows}"
    sql = (
        f"SELECT * FROM read_xlsx({quote_literal(str(path))}, sheet={quote_literal(sheet)}, "
        f"range={quote_literal(probe_range)}, header=false, all_varchar=true, stop_at_empty=false)"
    )
    return connection.execute(sql).fetchall()


def detect_header_row(grid: Sequence[Sequence[Any]]) -> Optional[int]:
    """1-based index of the most plausible header row in a probed grid.

    Heuristic: the first row with at least two non-empty cells whose following
    row — when there is one — also has at least two. A title row fails it (one
    cell) and a blank spacer fails it (zero cells), which covers the shape that
    broke the naive read: title, blank, header, data.
    """

    def width(row: Sequence[Any]) -> int:
        return sum(1 for cell in row if cell is not None and str(cell).strip() != "")

    for index, row in enumerate(grid):
        if width(row) < 2:
            continue
        following = grid[index + 1] if index + 1 < len(grid) else None
        if following is not None and width(following) < 2:
            continue
        return index + 1

    # Nothing looked like a header (single-column sheet, or data with no header
    # at all). Fall back to the first non-empty row so the load still produces
    # something the model can inspect and correct via ``header_row``.
    for index, row in enumerate(grid):
        if width(row) >= 1:
            return index + 1
    return None


# --------------------------------------------------------------- scan clauses


def _decodes(sample: bytes, encoding: str) -> bool:
    """Whether ``sample`` decodes cleanly, tolerating a truncated final character."""
    import codecs

    try:
        # Incremental, with ``final=False``: a fixed-size sample almost always
        # ends mid-character, and a one-shot decode would report that as a
        # failure for the very encoding that is correct.
        codecs.getincrementaldecoder(encoding)().decode(sample, False)
        return True
    except (UnicodeDecodeError, LookupError):
        return False


def detect_csv_encoding(path: Path, sample_bytes: int = 256 * 1024) -> str:
    """Best guess at a CSV's text encoding, as a name DuckDB's reader accepts.

    Needed because DuckDB assumes UTF-8 and *fails the whole read* on anything
    else — verified: a GB18030 file errors with ``CSV Error on Line: 1``. Excel's
    "Save as CSV" on a Chinese Windows writes GB18030, so that is not an exotic
    case, and preserving the bytes through upload (which the IDE now does) only
    gets them as far as a reader that cannot decode them.

    A BOM is decisive where present, and clean UTF-8 short-circuits. Beyond that
    ``charset_normalizer`` (already a runtime dependency) supplies a hint that is
    then *verified by decoding*, because on a short sample the CJK double-byte
    ranges overlap enough for it to mislabel freely — a GB18030 file comes back
    as ``cp949``. Trusting the label alone therefore fails the read outright.

    Discriminating between the CJK pages is still a heuristic, which is why the
    chosen encoding is reported on the result and can be overridden per call.
    When nothing decodes, the answer is ``utf-8`` rather than a single-byte page:
    Latin-1 accepts any byte sequence, so it would convert a miss into silent
    mojibake, whereas UTF-8 produces DuckDB's own actionable error.
    """
    try:
        with path.open("rb") as handle:
            sample = handle.read(sample_bytes)
    except OSError:  # pragma: no cover - caller has already stat'd the file
        return "utf-8"

    if sample.startswith(b"\xef\xbb\xbf"):
        return "utf-8"
    if sample.startswith((b"\xff\xfe", b"\xfe\xff")):
        return "utf-16"

    # Incremental, like the candidate loop below: a fixed-size sample almost
    # always ends mid-character, and a strict one-shot decode reports that as
    # "not UTF-8". The file then falls through to the CJK candidates, where
    # gb18030 accepts the bytes — so a plain UTF-8 file gets read as gb18030, or
    # worse, refused outright as Big5.
    if _decodes(sample, "utf-8"):
        return "utf-8"
    label = None
    guess = None
    try:
        from charset_normalizer import from_bytes

        best = from_bytes(sample).best()
        if best is not None and best.encoding:
            label = best.encoding
            guess = _ENCODING_ALIASES.get(label.replace("-", "_").lower())
    except ImportError:  # pragma: no cover - declared runtime dependency
        logger.debug("charset_normalizer unavailable for %s", path.name)

    # The label is a hint, not the answer. On a short sample the CJK double-byte
    # ranges overlap enough that detection mislabels freely — a GB18030 file
    # comes back as ``cp949`` — so whatever it says is confirmed by actually
    # decoding, and an unusable label falls through to the candidate list.
    # Only trust a "broken encoding" label on a sample big enough for the
    # detector to be confident. On a handful of bytes it will label almost
    # anything, and refusing a valid file on that basis is worse than the
    # mojibake the refusal exists to prevent — a file that short has no rows to
    # corrupt. A header plus one row is already well past this.
    if guess in _DUCKDB_BROKEN_CSV_ENCODINGS and len(sample) >= _MIN_SAMPLE_FOR_REFUSAL:
        raise DataFileError(
            f"{path.name} looks {guess}-encoded, which DuckDB's CSV reader corrupts "
            f"(it loses row boundaries after certain trailing bytes). Re-save the file as "
            f"UTF-8, or pass encoding={guess!r} explicitly to read it anyway."
        )

    ordered = [guess] if guess in _ENCODING_CANDIDATES else []
    ordered += [enc for enc in _ENCODING_CANDIDATES if enc != guess]
    for candidate in ordered:
        if _decodes(sample, candidate):
            if candidate != guess:
                logger.info("Encoding for %s reported as %s, reading as %s", path.name, label, candidate)
            return candidate

    # Nothing decoded — let DuckDB fail on utf-8 with its own actionable message
    # rather than forcing a single-byte page and returning mojibake.
    return "utf-8"


def _csv_scan(path: Path, encoding: Optional[str] = None) -> str:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    resolved = encoding or detect_csv_encoding(path)
    return (
        f"read_csv_auto({quote_literal(str(path))}, delim={quote_literal(delimiter)}, "
        f"encoding={quote_literal(resolved)})"
    )


def _parquet_scan(path: Path) -> str:
    return f"read_parquet({quote_literal(str(path))})"


def _json_scan(path: Path) -> str:
    return f"read_json_auto({quote_literal(str(path))})"


def _xlsx_scan(path: Path, sheet: str, header_row: int, max_row: int, max_column: int) -> Tuple[str, str]:
    """``(scan_expression, used_range)`` for one sheet, bounded to its real extent.

    ``stop_at_empty=false`` is stated rather than inferred. The range already
    ends at the last cell holding a value (:func:`sheet_bounds`), so there is no
    padding for it to protect against — and letting DuckDB infer it would
    truncate at the first interior blank row, silently dropping every group below
    a separator row, which hand-maintained exports use freely.
    """
    used_range = f"A{header_row}:{column_letter(max_column)}{max_row}"
    scan = (
        f"read_xlsx({quote_literal(str(path))}, sheet={quote_literal(sheet)}, "
        f"range={quote_literal(used_range)}, header=true, stop_at_empty=false)"
    )
    return scan, used_range


def enumerate_xls_sheets(path: Path) -> List[str]:
    """Sheet names of a legacy ``.xls`` workbook, in workbook order."""
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise DataFileError(
            "Reading legacy .xls requires the 'xlrd' package, which is not installed in this environment."
        ) from exc

    try:
        book = xlrd.open_workbook(str(path), on_demand=True)
    except Exception as exc:
        raise DataFileError(f"Failed to read legacy .xls file {path.name}: {exc}") from exc
    try:
        return list(book.sheet_names())
    finally:
        book.release_resources()


def convert_legacy_xls(path: Path, cache_dir: Path, sheet: Optional[str] = None) -> Path:
    """Convert one sheet of a legacy ``.xls`` to Parquet and return that path.

    DuckDB cannot read BIFF at all, so this is the only way in. ``sheet`` is not
    optional in spirit: ``pandas.read_excel`` defaults to the *first* worksheet,
    so a caller asking for a specific sheet of a multi-sheet workbook would
    silently receive a different one's data — wrong answers rather than an error.

    The conversion is cached next to the pod-local catalog and re-run whenever
    the source is newer, which is the one place in this module needing an
    explicit staleness check: every other format is read lazily and is therefore
    always current.
    """
    try:
        import pandas as pd
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise DataFileError("Reading legacy .xls requires 'pandas' and 'xlrd'.") from exc

    cache_dir.mkdir(parents=True, exist_ok=True)
    stat = path.stat()
    identity = _short_hash(str(path), sheet or "")
    cached = cache_dir / f"{path.stem}_{identity}_{int(stat.st_mtime_ns)}_{stat.st_size}.parquet"
    if cached.exists():
        return cached

    # Stale conversions of the same source+sheet (older mtime/size) are dead weight.
    for stale in cache_dir.glob(f"{path.stem}_{identity}_*.parquet"):
        try:
            stale.unlink()
        except OSError:  # pragma: no cover - best effort
            logger.debug("Could not remove stale xls conversion %s", stale)

    try:
        frame = pd.read_excel(path, engine="xlrd", sheet_name=sheet if sheet is not None else 0)
    except ImportError as exc:
        raise DataFileError(
            "Reading legacy .xls requires the 'xlrd' package, which is not installed in this environment."
        ) from exc
    except Exception as exc:
        raise DataFileError(f"Failed to read legacy .xls file {path.name}: {exc}") from exc

    frame.to_parquet(cached, index=False)
    return cached


# ------------------------------------------------------------------ profiling


_DATE_PROBE_BATCH = 40
"""Columns per text-date probe statement.

The probe's cost is quadratic in expressions per statement, not in total work:
1600 columns cost 23.8s as one statement and 0.57s in batches of 40 (DuckDB
1.5.2). 40 columns is 160 aggregates, measured at ~0.01s.
"""


def summarize_view(connection: Any, table: str) -> List[Dict[str, Any]]:
    """Per-column profile via DuckDB's ``SUMMARIZE``.

    Cheaper and more trustworthy than hand-rolling the same statistics in
    pandas, and it works on a lazy VIEW, so no materialisation is implied.
    """
    quoted = quote_identifier(table)
    cursor = connection.execute(f"SUMMARIZE {quoted}")
    names = [description[0] for description in cursor.description]
    keep = {
        "column_name",
        "column_type",
        "min",
        "max",
        "approx_unique",
        "avg",
        "count",
        "null_percentage",
    }
    profile: List[Dict[str, Any]] = []
    for row in cursor.fetchall():
        record = dict(zip(names, row))
        profile.append({key: _jsonable(value) for key, value in record.items() if key in keep})
    _annotate_text_dates(connection, table, profile)
    return profile


def _annotate_text_dates(connection: Any, table: str, profile: List[Dict[str, Any]]) -> None:
    """Mark VARCHAR columns that hold nothing but dates, with the cast to use.

    A spreadsheet column of dates typed as text stays VARCHAR: ``read_xlsx``
    converts a real date *cell* by its number format, but text is text. Every
    date function then fails to bind — ``strftime(order_date, '%Y-%m')`` reports
    a candidate-function error naming types, not the fix — and the model burns a
    round trip per attempt working out that a cast was needed.

    The cast is *verified*, not guessed from the column name or from SUMMARIZE's
    min/max: those are the lexical extremes, so a mixed column can show ISO ends
    and non-date values in between, and a hint that NULLs rows is worse than no
    hint. Only a column where every value parses gets annotated.

    Deliberately a hint and not a cast baked into the view: DuckDB re-expands a
    view's ``SELECT *`` per query, which is what makes a column added to the file
    later show up without reloading. Freezing an explicit projection to carry
    casts would trade that away, and it would buy little — TRY_CAST accepts only
    ISO-ish text (``8/31/2017`` and ``20170701`` both fail), which is exactly the
    text that already sorts and compares correctly as-is.

    Validity is tested against DATE, and TIMESTAMP is only *offered*. The two are
    not ordered the way they look: trailing whitespace after a bare date makes
    the TIMESTAMP cast NULL while the DATE cast still succeeds (DuckDB 1.5.2 —
    ``TRY_CAST('2017-07-01 ' AS TIMESTAMP)`` is NULL, ``AS DATE`` is not), and
    space-padded dates are ordinary in text cells and aligned CSV exports. Using
    TIMESTAMP as the gate silently withheld the hint from exactly those columns.

    Probed in batches because the cost is quadratic in the number of expressions
    per statement, and this runs inside ``exclusive_connection`` — the lock every
    concurrent ``execute_sql`` on the uploads catalog contends for. Measured on
    DuckDB 1.5.2 with a 1600-column view: 23.8s as one statement, 0.57s in
    batches of 40. Wide exports reach four-digit column counts and xlsx allows
    16384, so an unbatched probe would block every parallel tool call for the
    duration.
    """
    candidates = [
        column["column_name"] for column in profile if str(column.get("column_type", "")).upper() == "VARCHAR"
    ]
    if not candidates:
        return

    by_name = {column["column_name"]: column for column in profile}
    quoted_table = quote_identifier(table)
    for start in range(0, len(candidates), _DATE_PROBE_BATCH):
        batch = candidates[start : start + _DATE_PROBE_BATCH]
        selects = []
        for index, name in enumerate(batch):
            quoted = quote_identifier(name)
            selects.append(f"count({quoted}) AS n{index}")
            selects.append(
                f"count(*) FILTER (WHERE {quoted} IS NOT NULL AND TRY_CAST({quoted} AS DATE) IS NULL) AS bad_date{index}"
            )
            selects.append(
                f"count(*) FILTER (WHERE {quoted} IS NOT NULL AND TRY_CAST({quoted} AS TIMESTAMP) IS NULL) "
                f"AS bad_stamp{index}"
            )
            selects.append(
                f"count(*) FILTER (WHERE TRY_CAST({quoted} AS TIMESTAMP)::TIME <> TIME '00:00:00') AS clock{index}"
            )
        try:
            row = connection.execute(f"SELECT {', '.join(selects)} FROM {quoted_table}").fetchone()
        except Exception as exc:
            # A profile without hints is still a usable profile, so this stays
            # advisory: never fail a load over it.
            logger.debug("Could not probe %s for text date columns: %s", table, exc)
            return
        if row is None:
            return

        for index, name in enumerate(batch):
            non_null, unparsed, unparsed_stamp, with_clock = row[index * 4 : index * 4 + 4]
            if not non_null or unparsed:
                continue
            # A clock only earns the TIMESTAMP cast if *every* value survives it:
            # a column mixing timed values with space-padded bare dates would
            # otherwise be handed a cast that NULLs the padded ones. DATE is valid
            # for all of them, and losing a time is better than losing a row.
            target = "TIMESTAMP" if with_clock and not unparsed_stamp else "DATE"
            # Quoted unconditionally, like ``example_sql`` does: the hint exists to be
            # copied into a query, and a spreadsheet header with a space in it (or a
            # dash, or a reserved word) makes the bare form a parser error.
            expression = f"CAST({quote_identifier(name)} AS {target})"
            by_name[name]["cast_hint"] = f"text {target.lower()}s: {expression} to use date functions"


def preview_view(connection: Any, table: str, limit: int = PREVIEW_ROW_LIMIT) -> Tuple[List[str], List[List[Any]]]:
    quoted = quote_identifier(table)
    cursor = connection.execute(f"SELECT * FROM {quoted} LIMIT {int(limit)}")
    names = [description[0] for description in cursor.description]
    rows = [[_jsonable(value) for value in row] for row in cursor.fetchall()]
    return names, rows


def count_view(connection: Any, table: str) -> int:
    quoted = quote_identifier(table)
    row = connection.execute(f"SELECT count(*) FROM {quoted}").fetchone()
    return int(row[0]) if row else 0


def _jsonable(value: Any) -> Any:
    """Coerce DuckDB scalars into something a tool result can carry.

    ``SUMMARIZE`` returns ``null_percentage`` as ``Decimal`` and dates as
    ``date``/``datetime``; both break naive JSON serialisation downstream.
    """
    import datetime
    import decimal

    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, decimal.Decimal):
        return float(value)
    if isinstance(value, (datetime.date, datetime.datetime, datetime.time)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", "replace")
    return str(value)


# ----------------------------------------------------------------- the loader


def ownership_tag(rel_path: str, sheet: Optional[str]) -> str:
    """Marker stored as the object's COMMENT, identifying what a table came from.

    The catalog is the only bookkeeping this module keeps, and this tag is why it
    is enough. Without it, a reload cannot tell "this name is mine, replace it"
    from "this name belongs to a different file, pick another" — and guessing
    either way is broken: always replacing silently clobbers an unrelated
    upload, always suffixing makes reloads non-idempotent and grows the catalog
    without bound. The tag doubles as the provenance shown by ``describe_table``.
    """
    suffix = f"|sheet={sheet}" if sheet is not None else ""
    return f"datus:source={rel_path}{suffix}"


def registered_objects(connection: Any) -> Dict[str, Optional[str]]:
    """Every name in the catalog mapped to its ownership comment (``None`` if unset).

    Failures propagate. An empty catalog is a legitimate answer that this
    returns as ``{}`` — a fresh database answers both queries with zero rows,
    no error — so swallowing an exception into the same ``{}`` erases the only
    difference that matters to the caller. The SQL guard reads "no names" as
    "no table reference can be authorised" and refuses the query, which turns
    an unreadable catalog into a confident, wrong, and very actionable-looking
    "that table is not registered" about a table that is registered.
    """
    rows = connection.execute(
        "SELECT view_name, comment FROM duckdb_views() WHERE database_name != 'system' "
        "UNION ALL "
        "SELECT table_name, comment FROM duckdb_tables() WHERE database_name != 'system'"
    ).fetchall()
    return {row[0]: row[1] for row in rows}


def _names_owned_by_others(existing: Dict[str, Optional[str]], tag: str) -> List[str]:
    return [name for name, comment in existing.items() if comment != tag]


def _drop_existing(connection: Any, table: str) -> None:
    """Drop whatever currently owns ``table``, view or table.

    ``CREATE OR REPLACE`` only replaces an object of the *same* kind, and
    ``DROP TABLE IF EXISTS`` is not a no-op when a view holds the name — DuckDB
    raises "Existing object X is of type View, trying to drop type Table". So the
    kind has to be looked up first. This is reached whenever a reload flips
    ``materialize``, which is exactly when a user is iterating.
    """
    quoted = quote_identifier(table)
    literal = quote_literal(table)
    try:
        is_view = connection.execute(
            f"SELECT count(*) FROM duckdb_views() WHERE view_name = {literal} AND database_name != 'system'"
        ).fetchone()
        if is_view and is_view[0]:
            connection.execute(f"DROP VIEW IF EXISTS {quoted}")
            return
        is_table = connection.execute(
            f"SELECT count(*) FROM duckdb_tables() WHERE table_name = {literal} AND database_name != 'system'"
        ).fetchone()
        if is_table and is_table[0]:
            connection.execute(f"DROP TABLE IF EXISTS {quoted}")
    except Exception as exc:  # pragma: no cover - catalog probe is best effort
        logger.debug("Could not probe existing object %s: %s", table, exc)


def _create_view(connection: Any, table: str, scan: str, materialize: bool, tag: str) -> None:
    _drop_existing(connection, table)
    quoted = quote_identifier(table)
    if materialize:
        # Snapshot semantics, chosen explicitly by the caller: worth it when the
        # source is large enough that re-scanning it per query dominates, at the
        # cost of no longer tracking edits to the file.
        connection.execute(f"CREATE TABLE {quoted} AS SELECT * FROM {scan}")
        kind = "TABLE"
    else:
        connection.execute(f"CREATE VIEW {quoted} AS SELECT * FROM {scan}")
        kind = "VIEW"
    # Set after creation, not via CREATE OR REPLACE: replacing an object drops
    # its comment, so the tag has to be (re)applied on every load.
    connection.execute(f"COMMENT ON {kind} {quoted} IS {quote_literal(tag)}")


def _finish_table(
    connection: Any,
    *,
    table: str,
    source_file: str,
    sheet: Optional[str],
    header_row: Optional[int],
    used_range: Optional[str],
    materialized: bool,
    encoding: Optional[str] = None,
) -> LoadedTable:
    preview_columns, preview_rows = preview_view(connection, table)
    return LoadedTable(
        table=table,
        source_file=source_file,
        sheet=sheet,
        header_row=header_row,
        used_range=used_range,
        encoding=encoding,
        row_count=count_view(connection, table),
        columns=summarize_view(connection, table),
        preview_columns=preview_columns,
        preview_rows=preview_rows,
        materialized=materialized,
    )


def inspect_file(path: Path, *, connection: Any, sheet: Optional[str] = None) -> Dict[str, Any]:
    """Describe a file without creating anything.

    The point is the *un-interpreted* grid for spreadsheets: when the header
    guess is wrong, the model needs to see the raw upper-left cells to work out
    the right ``header_row``, and a parsed preview cannot show that.
    """
    suffix = path.suffix.lower()
    result: Dict[str, Any] = {"file": path.name, "format": suffix.lstrip(".")}

    if suffix in _EXCEL_SUFFIXES:
        sheets = enumerate_sheets(path)
        result["sheets"] = sheets
        targets = [sheet] if sheet else sheets
        grids = []
        for name in targets:
            if name not in sheets:
                raise DataFileError(f"Sheet {name!r} not found in {path.name}. Available: {sheets}")
            max_row, max_column = sheet_bounds(path, name)
            if max_row < 1 or max_column < 1:
                grids.append({"sheet": name, "empty": True})
                continue
            grid = _probe_grid(connection, path, name, max_row, max_column)
            grids.append(
                {
                    "sheet": name,
                    "used_rows": max_row,
                    "used_columns": max_column,
                    "detected_header_row": detect_header_row(grid),
                    "raw_rows": [[_jsonable(cell) for cell in row] for row in grid],
                }
            )
        result["sheet_details"] = grids
        return result

    if suffix in _CSV_SUFFIXES:
        detected = detect_csv_encoding(path)
        result["encoding"] = detected
        scan = _csv_scan(path, detected)
    elif suffix in _PARQUET_SUFFIXES:
        scan = _parquet_scan(path)
    elif suffix in _JSON_SUFFIXES:
        scan = _json_scan(path)
    elif suffix in _LEGACY_EXCEL_SUFFIXES:
        result["note"] = "Legacy .xls is converted to Parquet on load; run without inspect_only to query it."
        return result
    else:
        raise DataFileError(_unsupported_message(path))

    cursor = connection.execute(f"SELECT * FROM {scan} LIMIT {PREVIEW_ROW_LIMIT}")
    result["preview_columns"] = [description[0] for description in cursor.description]
    result["preview_rows"] = [[_jsonable(value) for value in row] for row in cursor.fetchall()]
    return result


def _unsupported_message(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in _UNSUPPORTED_SUFFIXES:
        return f"{suffix} files are not supported. Re-save {path.name} as .xlsx or .csv and upload that instead."
    return (
        f"Unsupported data file format {suffix or '(none)'} for {path.name}. "
        f"Supported: {', '.join(sorted(SUPPORTED_SUFFIXES))}."
    )


def load_file(
    path: Path,
    rel_path: str,
    *,
    connection: Any,
    conversion_cache_dir: Path,
    sheet: Optional[str] = None,
    header_row: Optional[int] = None,
    materialize: bool = False,
    encoding: Optional[str] = None,
    existing_objects: Optional[Dict[str, Optional[str]]] = None,
) -> Tuple[List[LoadedTable], List[SkippedSheet]]:
    """Create one VIEW (or table, when ``materialize``) per loadable unit.

    A spreadsheet's unit is a sheet; every other format is a single unit. Sheets
    that cannot be read are collected into the skipped list rather than failing
    the whole load — an empty "notes" tab is normal in real workbooks and must
    not cost the user the sheets that do have data.
    """
    suffix = path.suffix.lower()
    existing = dict(existing_objects or {})
    loaded: List[LoadedTable] = []
    skipped: List[SkippedSheet] = []

    if suffix in _EXCEL_SUFFIXES:
        sheets = enumerate_sheets(path)
        if sheet is not None:
            if sheet not in sheets:
                raise DataFileError(f"Sheet {sheet!r} not found in {path.name}. Available: {sheets}")
            targets = [(sheets.index(sheet), sheet)]
        else:
            targets = list(enumerate(sheets))

        for index, name in targets:
            try:
                max_row, max_column = sheet_bounds(path, name)
                if max_row < 1 or max_column < 1:
                    skipped.append(SkippedSheet(name, "sheet is empty"))
                    continue
                resolved_header = header_row
                if resolved_header is not None and resolved_header < 1:
                    skipped.append(SkippedSheet(name, f"header_row must be 1 or greater, got {resolved_header}"))
                    continue
                if resolved_header is None:
                    grid = _probe_grid(connection, path, name, max_row, max_column)
                    resolved_header = detect_header_row(grid)
                if resolved_header is None:
                    skipped.append(SkippedSheet(name, "no header row could be located"))
                    continue
                if resolved_header >= max_row + 1:
                    skipped.append(SkippedSheet(name, "header row is past the end of the sheet"))
                    continue

                tag = ownership_tag(rel_path, name)
                table = build_table_name(
                    rel_path=rel_path,
                    sheet=name,
                    sheet_index=index,
                    taken=_names_owned_by_others(existing, tag),
                )
                scan, used_range = _xlsx_scan(path, name, resolved_header, max_row, max_column)
                _create_view(connection, table, scan, materialize, tag)
                existing[table] = tag
                loaded.append(
                    _finish_table(
                        connection,
                        table=table,
                        source_file=rel_path,
                        sheet=name,
                        header_row=resolved_header,
                        used_range=used_range,
                        materialized=materialize,
                    )
                )
            except DataFileError as exc:
                skipped.append(SkippedSheet(name, str(exc)))
            except Exception as exc:
                # DuckDB raises for an empty sheet ("No rows found in xlsx
                # file") among other per-sheet problems. One bad tab must not
                # sink the workbook.
                logger.info("Skipping sheet %s of %s: %s", name, rel_path, exc)
                # An exception with an empty message yields [] from splitlines(),
                # so indexing it would make this handler the thing that fails.
                reason = (str(exc).splitlines() or [repr(exc)])[0]
                skipped.append(SkippedSheet(name, reason))

        if not loaded and skipped:
            reasons = "; ".join(f"{item.sheet}: {item.reason}" for item in skipped)
            raise DataFileError(f"No readable sheet in {path.name} ({reasons})")
        return loaded, skipped

    resolved_encoding: Optional[str] = None
    if suffix in _CSV_SUFFIXES:
        resolved_encoding = encoding or detect_csv_encoding(path)
        scan = _csv_scan(path, resolved_encoding)
    elif suffix in _PARQUET_SUFFIXES:
        scan = _parquet_scan(path)
    elif suffix in _JSON_SUFFIXES:
        scan = _json_scan(path)
    elif suffix in _LEGACY_EXCEL_SUFFIXES:
        return _load_legacy_xls(
            path,
            rel_path,
            connection=connection,
            conversion_cache_dir=conversion_cache_dir,
            sheet=sheet,
            materialize=materialize,
            existing=existing,
        )
    else:
        raise DataFileError(_unsupported_message(path))

    tag = ownership_tag(rel_path, None)
    table = build_table_name(
        rel_path=rel_path,
        sheet=None,
        sheet_index=0,
        taken=_names_owned_by_others(existing, tag),
    )
    _create_view(connection, table, scan, materialize, tag)
    loaded.append(
        _finish_table(
            connection,
            table=table,
            source_file=rel_path,
            sheet=None,
            header_row=None,
            used_range=None,
            materialized=materialize,
            encoding=resolved_encoding,
        )
    )
    return loaded, skipped


def _load_legacy_xls(
    path: Path,
    rel_path: str,
    *,
    connection: Any,
    conversion_cache_dir: Path,
    sheet: Optional[str],
    materialize: bool,
    existing: Dict[str, Optional[str]],
) -> Tuple[List[LoadedTable], List[SkippedSheet]]:
    """One table per sheet, matching the ``.xlsx`` path.

    A legacy workbook has sheets like any other, so treating the file as a single
    unit made ``sheet`` a no-op — a caller asking for the second sheet silently
    got the first one's data. Each sheet converts to its own Parquet file.
    """
    sheets = enumerate_xls_sheets(path)
    if sheet is not None:
        if sheet not in sheets:
            raise DataFileError(f"Sheet {sheet!r} not found in {path.name}. Available: {sheets}")
        targets = [(sheets.index(sheet), sheet)]
    else:
        targets = list(enumerate(sheets))

    loaded: List[LoadedTable] = []
    skipped: List[SkippedSheet] = []
    for index, name in targets:
        try:
            parquet = convert_legacy_xls(path, conversion_cache_dir, sheet=name)
            tag = ownership_tag(rel_path, name)
            table = build_table_name(
                rel_path=rel_path,
                sheet=name if len(sheets) > 1 else None,
                sheet_index=index,
                taken=_names_owned_by_others(existing, tag),
            )
            _create_view(connection, table, _parquet_scan(parquet), materialize, tag)
            existing[table] = tag
            loaded.append(
                _finish_table(
                    connection,
                    table=table,
                    source_file=rel_path,
                    sheet=name,
                    header_row=None,
                    used_range=None,
                    materialized=materialize,
                )
            )
        except DataFileError as exc:
            skipped.append(SkippedSheet(name, str(exc)))
        except Exception as exc:
            logger.info("Skipping sheet %s of %s: %s", name, rel_path, exc)
            skipped.append(SkippedSheet(name, (str(exc).splitlines() or [repr(exc)])[0]))

    if not loaded and skipped:
        reasons = "; ".join(f"{item.sheet}: {item.reason}" for item in skipped)
        raise DataFileError(f"No readable sheet in {path.name} ({reasons})")
    return loaded, skipped


def default_conversion_cache_dir(db_path: str) -> Path:
    """Where ``.xls`` → Parquet conversions live: next to the DuckDB catalog.

    Same lifetime and same locality as the catalog itself, so a pod losing one
    loses the other and both rebuild together.
    """
    parent = os.path.dirname(db_path)
    if not parent:
        # An in-memory or unnamed catalog has no directory to sit beside, and the
        # process working directory is not ours to write into.
        return Path(tempfile.gettempdir()) / "datus_xls_conversions"
    return Path(parent) / "_conversions"
